from openpyxl import load_workbook
import csv




def getdata():
    wb=load_workbook(filename='/home/karan/Downloads/doc/Students_List.xlsx')
    sheet_range=wb['Regular']
    with open('studentDetails.csv','w') as csv_file:
        
        field_names=['uid','name']
        csv_writer=csv.DictWriter(csv_file,field_names)
        csv_writer.writeheader()
        for i in sheet_range.iter_rows(min_row=1,max_col=3 ):
              dic={'uid':i[1].value,'name':i[2].value}
              csv_writer.writerow(dic)




def search(name):
    with open('studentDetails.csv','r') as csv_file:
        csv_reader=csv.DictReader(csv_file)
        i=1
        for line in csv_reader:
            if name.upper() in line['name'] or name in line['name']:
                print(f"{i}. UID of {line['name']}: {line['uid']}")
                i+=1

def get_input():
    name=input('Enter name of student:')
    return name



while True:
  search(get_input())
  choice=input('Want to search more(y/n):')
  if choice != 'y' and choice !='Y':
      break








#Uncomment it to extract data from excel file
#getdata()
